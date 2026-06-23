"""Excel (.xlsx) file parser implementation."""

from pathlib import Path

from lightrag_core.ingestion.parser import BaseParser


class XlsxParser(BaseParser):
    """Parser for Excel .xlsx files.

    Iterates all sheets, converts each row to "column: value" format.
    """

    SUPPORTED_EXTENSIONS = {".xlsx"}

    def parse(self, file_path: str | Path) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("openpyxl required for .xlsx parsing: pip install openpyxl")

        path = Path(file_path)
        wb = load_workbook(path, read_only=True, data_only=True)
        lines: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    headers = [str(c) if c is not None else f"col{i}" for i, c in enumerate(row)]
                    continue

                if all(v is None or str(v).strip() == "" for v in row):
                    continue

                parts = []
                for i in range(min(len(headers), len(row))):
                    value = row[i]
                    if value is not None and str(value).strip():
                        parts.append(f"{headers[i]}: {value}")

                if parts:
                    lines.append("; ".join(parts))

        wb.close()
        return "\n".join(lines)

    def supports(self, file_path: str | Path) -> bool:
        return Path(file_path).suffix.lower() in self.SUPPORTED_EXTENSIONS
